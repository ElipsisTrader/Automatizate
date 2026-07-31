from __future__ import annotations

import hashlib
import io
import tempfile
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_NAME = "Rendición Inteligente"
APP_VERSION = "0.1.0"

COLUMNAS = [
    "archivo",
    "tipo_documento",
    "rut_emisor",
    "razon_social",
    "fecha",
    "folio",
    "monto_neto",
    "monto_iva",
    "monto_total",
    "glosa",
    "estado",
    "observacion",
]

ALIAS_CAMPOS = {
    "nombre_archivo": "archivo",
    "documento": "archivo",
    "tipo": "tipo_documento",
    "rut": "rut_emisor",
    "emisor_rut": "rut_emisor",
    "comercio": "razon_social",
    "empresa": "razon_social",
    "fecha_emision": "fecha",
    "numero_documento": "folio",
    "numero_folio": "folio",
    "neto": "monto_neto",
    "iva": "monto_iva",
    "total": "monto_total",
    "glosa_sugerida": "glosa",
    "descripcion": "glosa",
    "status": "estado",
    "motivo_revision": "observacion",
}


def configurar_pagina() -> None:
    st.set_page_config(
        page_title=f"{APP_NAME} | Demo",
        page_icon="🧾",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 1.8rem;
            padding-bottom: 3rem;
            max-width: 1450px;
        }
        .hero {
            padding: 1.45rem 1.6rem;
            border: 1px solid rgba(128,128,128,.22);
            border-radius: 18px;
            background: linear-gradient(135deg, rgba(26,115,232,.10), rgba(0,184,148,.08));
            margin-bottom: 1.2rem;
        }
        .hero h1 {
            margin: 0 0 .35rem 0;
            font-size: 2rem;
        }
        .hero p {
            margin: 0;
            opacity: .82;
            font-size: 1.02rem;
        }
        .demo-warning {
            border-left: 5px solid #f39c12;
            background: rgba(243,156,18,.10);
            padding: .85rem 1rem;
            border-radius: 8px;
            margin: .6rem 0 1rem 0;
        }
        div[data-testid="stMetric"] {
            border: 1px solid rgba(128,128,128,.20);
            padding: 1rem;
            border-radius: 14px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def inicializar_estado() -> None:
    if "resultados" not in st.session_state:
        st.session_state.resultados = pd.DataFrame(columns=COLUMNAS)
    if "procesamiento_realizado" not in st.session_state:
        st.session_state.procesamiento_realizado = False


def normalizar_resultado(resultado: Any, nombre_archivo: str) -> dict[str, Any]:
    """Adapta la salida del motor real a la estructura visual."""
    if resultado is None:
        raise ValueError("El motor no devolvió información.")

    if hasattr(resultado, "model_dump"):
        datos = resultado.model_dump()
    elif isinstance(resultado, dict):
        datos = dict(resultado)
    elif hasattr(resultado, "__dict__"):
        datos = dict(vars(resultado))
    else:
        raise TypeError(
            "La salida del motor debe ser un dict, un modelo Pydantic "
            "o un objeto con atributos."
        )

    datos_normalizados: dict[str, Any] = {}
    for clave, valor in datos.items():
        clave_destino = ALIAS_CAMPOS.get(str(clave), str(clave))
        datos_normalizados[clave_destino] = valor

    valores_por_defecto = {
        "archivo": nombre_archivo,
        "tipo_documento": "Sin identificar",
        "rut_emisor": "",
        "razon_social": "",
        "fecha": "",
        "folio": "",
        "monto_neto": 0,
        "monto_iva": 0,
        "monto_total": 0,
        "glosa": "",
        "estado": "Procesado",
        "observacion": "",
    }
    for clave, valor in valores_por_defecto.items():
        datos_normalizados.setdefault(clave, valor)

    return {columna: datos_normalizados.get(columna, "") for columna in COLUMNAS}


def procesar_con_motor_real(ruta_archivo: str) -> dict[str, Any]:
    """
    PUNTO DE INTEGRACIÓN CON EL CÓDIGO REAL.

    Tu colega debe reemplazar el contenido de esta función por la llamada
    a su motor actual. Ejemplo:

        from motor_actual import procesar_documento
        return procesar_documento(ruta_archivo)
    """
    raise NotImplementedError(
        "El motor real todavía no está conectado a esta interfaz. "
        "Activa el modo demostración o integra aquí el código de extracción."
    )


def crear_resultado_demo(archivo, indice: int) -> dict[str, Any]:
    """Genera datos simulados. No analiza el contenido real."""
    contenido = archivo.getvalue()
    huella = hashlib.sha256(contenido + archivo.name.encode("utf-8")).hexdigest()
    semilla = int(huella[:10], 16)

    razones = [
        ("76.123.456-7", "Servicios Comerciales Andinos SpA"),
        ("77.654.321-K", "Distribuidora Central Ltda."),
        ("96.888.777-5", "Estación de Servicio Cordillera"),
        ("78.222.333-4", "Insumos Corporativos del Sur"),
        ("76.901.234-2", "Alimentos y Servicios Urbanos"),
    ]
    rut, razon = razones[semilla % len(razones)]

    extension = Path(archivo.name).suffix.lower()
    tipo = "Factura" if extension == ".pdf" or semilla % 3 == 0 else "Boleta"
    total = 8_000 + (semilla % 180_000)
    neto = round(total / 1.19)
    iva = total - neto

    fecha_demo = date.today() - timedelta(days=semilla % 25)
    requiere_revision = indice % 5 == 4 or "borros" in archivo.name.lower()

    if requiere_revision:
        estado = "Requiere revisión"
        observacion = "Dato de ejemplo: revisar legibilidad o campo incompleto."
    else:
        estado = "Procesado"
        observacion = ""

    return {
        "archivo": archivo.name,
        "tipo_documento": tipo,
        "rut_emisor": rut,
        "razon_social": razon,
        "fecha": fecha_demo.isoformat(),
        "folio": str(100_000 + semilla % 899_999),
        "monto_neto": neto,
        "monto_iva": iva,
        "monto_total": total,
        "glosa": "Gasto operativo de demostración",
        "estado": estado,
        "observacion": observacion,
    }


def procesar_archivos(archivos, modo_demo: bool) -> pd.DataFrame:
    registros: list[dict[str, Any]] = []
    barra = st.progress(0, text="Preparando documentos…")
    total_archivos = len(archivos)

    with tempfile.TemporaryDirectory(prefix="rendiciones_") as carpeta_temporal:
        carpeta = Path(carpeta_temporal)

        for indice, archivo in enumerate(archivos):
            barra.progress(
                indice / total_archivos,
                text=f"Procesando {indice + 1} de {total_archivos}: {archivo.name}",
            )

            ruta = carpeta / Path(archivo.name).name
            ruta.write_bytes(archivo.getvalue())

            try:
                if modo_demo:
                    resultado = crear_resultado_demo(archivo, indice)
                else:
                    resultado_motor = procesar_con_motor_real(str(ruta))
                    resultado = normalizar_resultado(resultado_motor, archivo.name)
                registros.append(resultado)
            except Exception as error:
                registros.append(
                    {
                        "archivo": archivo.name,
                        "tipo_documento": "Sin identificar",
                        "rut_emisor": "",
                        "razon_social": "",
                        "fecha": "",
                        "folio": "",
                        "monto_neto": 0,
                        "monto_iva": 0,
                        "monto_total": 0,
                        "glosa": "",
                        "estado": "Error",
                        "observacion": str(error),
                    }
                )

    barra.progress(1.0, text="Procesamiento finalizado")
    return pd.DataFrame(registros, columns=COLUMNAS)


def crear_excel(df: pd.DataFrame) -> bytes:
    salida = io.BytesIO()

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Rendiciones")
        hoja = writer.sheets["Rendiciones"]
        hoja.freeze_panes = "A2"
        hoja.auto_filter.ref = hoja.dimensions

        relleno = PatternFill("solid", fgColor="17365D")
        fuente = Font(color="FFFFFF", bold=True)

        for celda in hoja[1]:
            celda.fill = relleno
            celda.font = fuente
            celda.alignment = Alignment(horizontal="center", vertical="center")

        for indice_columna, nombre_columna in enumerate(df.columns, start=1):
            valores = [str(nombre_columna)]
            valores.extend(str(valor) for valor in df[nombre_columna].fillna("").tolist())
            ancho = min(max(len(valor) for valor in valores) + 2, 42)
            hoja.column_dimensions[get_column_letter(indice_columna)].width = ancho

        for columna in ["monto_neto", "monto_iva", "monto_total"]:
            numero_columna = df.columns.get_loc(columna) + 1
            for fila in range(2, len(df) + 2):
                hoja.cell(fila, numero_columna).number_format = '$#,##0'

    salida.seek(0)
    return salida.getvalue()


def mostrar_vista_previa(archivos) -> None:
    with st.expander("Ver archivos seleccionados", expanded=False):
        for archivo in archivos:
            extension = Path(archivo.name).suffix.lower()
            col1, col2 = st.columns([1, 3])

            with col1:
                st.write(f"**{archivo.name}**")
                st.caption(f"{len(archivo.getvalue()) / 1024:.1f} KB")

            with col2:
                if extension in {".jpg", ".jpeg", ".png"}:
                    st.image(archivo.getvalue(), width=280)
                else:
                    st.info("PDF cargado. La vista previa se habilitará más adelante.")


def mostrar_metricas(df: pd.DataFrame) -> None:
    total_documentos = len(df)
    procesados = int((df["estado"] == "Procesado").sum()) if total_documentos else 0
    revision = int((df["estado"] == "Requiere revisión").sum()) if total_documentos else 0
    errores = int((df["estado"] == "Error").sum()) if total_documentos else 0
    monto_total = pd.to_numeric(df["monto_total"], errors="coerce").fillna(0).sum()

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Documentos", total_documentos)
    c2.metric("Procesados", procesados)
    c3.metric("En revisión", revision)
    c4.metric("Errores", errores)
    c5.metric("Monto total", f"${monto_total:,.0f}".replace(",", "."))


def mostrar_resultados() -> None:
    df = st.session_state.resultados

    if df.empty:
        st.info("Todavía no hay resultados. Carga documentos y presiona «Procesar documentos».")
        return

    mostrar_metricas(df)
    st.markdown("### Resultados extraídos")
    st.caption("Puedes corregir los campos antes de descargar la planilla.")

    configuracion_columnas = {
        "archivo": st.column_config.TextColumn("Archivo", disabled=True),
        "tipo_documento": st.column_config.SelectboxColumn(
            "Tipo", options=["Boleta", "Factura", "Voucher", "Sin identificar"]
        ),
        "rut_emisor": st.column_config.TextColumn("RUT emisor"),
        "razon_social": st.column_config.TextColumn("Razón social"),
        "fecha": st.column_config.TextColumn("Fecha"),
        "folio": st.column_config.TextColumn("Folio"),
        "monto_neto": st.column_config.NumberColumn("Neto", format="$ %d"),
        "monto_iva": st.column_config.NumberColumn("IVA", format="$ %d"),
        "monto_total": st.column_config.NumberColumn("Total", format="$ %d"),
        "glosa": st.column_config.TextColumn("Glosa"),
        "estado": st.column_config.SelectboxColumn(
            "Estado", options=["Procesado", "Requiere revisión", "Error"]
        ),
        "observacion": st.column_config.TextColumn("Observación"),
    }

    df_editado = st.data_editor(
        df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config=configuracion_columnas,
        key="editor_resultados",
    )
    st.session_state.resultados = df_editado

    excel = crear_excel(df_editado)
    csv = df_editado.to_csv(index=False).encode("utf-8-sig")

    col_excel, col_csv, col_limpiar = st.columns([1, 1, 2])

    with col_excel:
        st.download_button(
            "Descargar Excel",
            data=excel,
            file_name="rendiciones_procesadas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_csv:
        st.download_button(
            "Descargar CSV",
            data=csv,
            file_name="rendiciones_procesadas.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col_limpiar:
        if st.button("Limpiar resultados", type="secondary"):
            st.session_state.resultados = pd.DataFrame(columns=COLUMNAS)
            st.session_state.procesamiento_realizado = False
            st.rerun()


def render_sidebar() -> bool:
    st.sidebar.title("Configuración")
    st.sidebar.caption(f"Demo v{APP_VERSION}")

    modo_demo = st.sidebar.toggle(
        "Modo demostración",
        value=True,
        help="Genera datos simulados. Desactívalo cuando el motor real esté conectado.",
    )

    if modo_demo:
        st.sidebar.warning("Los datos serán simulados. No se analizará el contenido real.")
    else:
        st.sidebar.success("Modo motor real seleccionado.")

    st.sidebar.divider()
    st.sidebar.markdown("**Estado del proyecto**")
    st.sidebar.write("✅ Motor Python existente")
    st.sidebar.write("✅ Interfaz visual")
    st.sidebar.write("⬜ Conectar motor a la interfaz")
    st.sidebar.write("⬜ Reglas y alertas reales")
    st.sidebar.write("⬜ WhatsApp")
    st.sidebar.write("⬜ Integración ERP")

    return modo_demo


def main() -> None:
    configurar_pagina()
    inicializar_estado()
    modo_demo = render_sidebar()

    st.markdown(
        """
        <div class="hero">
            <h1>🧾 Rendición Inteligente</h1>
            <p>Carga boletas y facturas, procesa la información y descarga una planilla consolidada.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if modo_demo:
        st.markdown(
            """
            <div class="demo-warning">
                <strong>Modo demostración activo:</strong>
                esta versión permite presentar el flujo visual, pero los datos mostrados son simulados.
            </div>
            """,
            unsafe_allow_html=True,
        )

    tab_procesar, tab_resultados, tab_integracion = st.tabs(
        ["1. Cargar y procesar", "2. Resultados", "3. Integración técnica"]
    )

    with tab_procesar:
        st.subheader("Carga de documentos")
        st.write("Selecciona varios documentos PDF, JPG, JPEG o PNG.")

        archivos = st.file_uploader(
            "Arrastra los archivos aquí o selecciónalos desde tu computador",
            type=["pdf", "jpg", "jpeg", "png"],
            accept_multiple_files=True,
        )

        if archivos:
            st.success(f"{len(archivos)} documento(s) preparado(s).")
            mostrar_vista_previa(archivos)

            if st.button("Procesar documentos", type="primary", use_container_width=True):
                with st.spinner("Ejecutando el flujo de procesamiento…"):
                    st.session_state.resultados = procesar_archivos(archivos, modo_demo)
                    st.session_state.procesamiento_realizado = True
                st.success("Proceso completado. Abre la pestaña «Resultados».")
        else:
            st.info("Carga al menos un documento para continuar.")

    with tab_resultados:
        mostrar_resultados()

    with tab_integracion:
        st.subheader("Cómo conectar el código real")
        st.write(
            "Tu colega debe modificar la función `procesar_con_motor_real(ruta_archivo)` "
            "dentro de este mismo archivo."
        )
        st.code(
            """def procesar_con_motor_real(ruta_archivo: str):
    from motor_actual import procesar_documento
    return procesar_documento(ruta_archivo)""",
            language="python",
        )
        st.markdown(
            """
            El motor puede devolver un diccionario o un modelo Pydantic con campos como:
            `tipo_documento`, `rut_emisor`, `razon_social`, `fecha_emision`, `folio`,
            `monto_neto`, `iva`, `monto_total` y `glosa_sugerida`.
            """
        )

    st.divider()
    st.caption(
        "Prototipo de demostración. Los resultados financieros deben mantenerse "
        "bajo revisión humana mientras el sistema se valida."
    )


if __name__ == "__main__":
    main()
